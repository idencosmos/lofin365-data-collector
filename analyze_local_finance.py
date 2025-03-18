import pandas as pd
import sqlite3
from datetime import datetime
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference, BarChart
import os
from config import config
import gc  # 메모리 관리를 위한 가비지 컬렉션 모듈

#################################################
# 지방재정 데이터 분석 클래스
# - 지방재정 데이터베이스 연결 및 분석 수행
# - 분석 결과 엑셀 보고서 생성
#################################################
class LocalFinanceAnalyzer:
    def __init__(self, db_path=config.db_name):
        """
        지방재정 분석기 초기화
        
        Args:
            db_path (str): 데이터베이스 파일 경로
        """
        self.db_path = db_path
        self.conn = None
        self.chunk_size = 50000
        
    #########################################
    # 1. 데이터베이스 연결 관리 메서드
    #########################################
    def connect_db(self):
        """데이터베이스 연결 설정"""
        if not self.conn:
            self.conn = sqlite3.connect(self.db_path)
        return self.conn
        
    def close_db(self):
        """데이터베이스 연결 종료"""
        if self.conn:
            self.conn.close()
            self.conn = None

    #########################################
    # 2. 개선된 분석 쿼리 메서드
    #########################################
    def get_annual_budget(self, year):
        """연간 총 예산: 12월 31일 기준 사업별 고유 예산액 합산"""
        conn = self.connect_db()
        query = f"""
        SELECT SUM(bdg_cash_amt) as annual_budget
        FROM yearly_data_{year}
        WHERE strftime('%m', exe_ymd) = '12'
        """
        result = pd.read_sql_query(query, conn)
        return result['annual_budget'].iloc[0] if not result.empty else 0

    def calculate_monthly_execution(self, year):
        """월별 집행 현황: 해당 월의 예산 사용"""
        conn = self.connect_db()
        annual_budget = self.get_annual_budget(year)
        query = f"""
        WITH monthly_data AS (
            SELECT 
                strftime('%m', exe_ymd) as month,
                SUM(bdg_cash_amt) as monthly_budget,  -- 해당 월의 예산
                SUM(ep_amt) as cumulative_execution,
                SUM(bdg_ntep) as national_budget,
                SUM(capep) as provincial_budget,
                SUM(sggep) as municipal_budget,
                SUM(etc_amt) as other_budget,
                SUM(cpl_amt) as allocated_budget  -- 편성액 추가
            FROM yearly_data_{year}
            GROUP BY month
        ),
        monthly_actual AS (
            SELECT 
                month,
                monthly_budget,
                cumulative_execution,
                national_budget,
                provincial_budget,
                municipal_budget,
                other_budget,
                allocated_budget,
                cumulative_execution - LAG(cumulative_execution, 1, 0) OVER (ORDER BY month) as monthly_execution
            FROM monthly_data
        )
        SELECT 
            month,
            monthly_budget as total_budget,
            monthly_execution as execution_amount,
            CASE WHEN {annual_budget} > 0 THEN monthly_execution / {annual_budget} ELSE 0 END as execution_rate,
            national_budget,
            provincial_budget,
            municipal_budget,
            other_budget,
            allocated_budget
        FROM monthly_actual
        ORDER BY month
        """
        return pd.read_sql_query(query, conn)

    def analyze_execution_by_month(self, year):
        """월별 예산 집행 현황 분석"""
        conn = self.connect_db()
        annual_budget = self.get_annual_budget(year)
        query = f"""
        WITH monthly_execution AS (
            SELECT 
                strftime('%m', exe_ymd) as month,
                SUM(bdg_cash_amt) as total_budget,  -- 해당 월의 예산
                SUM(ep_amt) as total_execution,
                SUM(bdg_ntep) as national_budget,
                SUM(capep) as provincial_budget,
                SUM(sggep) as municipal_budget,
                SUM(etc_amt) as other_budget,
                SUM(cpl_amt) as allocated_budget,  -- 편성액 추가
                COUNT(*) as record_count
            FROM yearly_data_{year}
            GROUP BY month
        )
        SELECT 
            month,
            total_budget,
            total_execution,
            CASE WHEN {annual_budget} > 0 THEN total_execution / {annual_budget} ELSE 0 END as cumulative_execution_rate,
            national_budget,
            provincial_budget,
            municipal_budget,
            other_budget,
            allocated_budget,
            record_count
        FROM monthly_execution
        ORDER BY month
        """
        return pd.read_sql_query(query, conn)

    def analyze_seasonality(self, year):
        """계절성 분석 (중복 제거 후 통합)"""
        conn = self.connect_db()
        query = f"""
        WITH monthly_data AS (
            SELECT 
                strftime('%m', exe_ymd) as month,
                SUM(ep_amt) as cumulative_execution,
                LAG(SUM(ep_amt)) OVER (ORDER BY strftime('%m', exe_ymd)) as prev_execution
            FROM yearly_data_{year}
            GROUP BY month
        ),
        monthly_actual AS (
            SELECT 
                month,
                CASE WHEN prev_execution IS NULL THEN cumulative_execution
                     ELSE cumulative_execution - prev_execution END as actual_execution
            FROM monthly_data
        ),
        quarterly_data AS (
            SELECT 
                CASE 
                    WHEN CAST(month AS INTEGER) BETWEEN 1 AND 3 THEN '1'
                    WHEN CAST(month AS INTEGER) BETWEEN 4 AND 6 THEN '2'
                    WHEN CAST(month AS INTEGER) BETWEEN 7 AND 9 THEN '3'
                    ELSE '4'
                END as quarter,
                SUM(actual_execution) as quarterly_execution
            FROM monthly_actual
            GROUP BY quarter
        )
        SELECT 
            quarter,
            quarterly_execution,
            quarterly_execution / SUM(quarterly_execution) OVER () * 100 as quarter_ratio
        FROM quarterly_data
        ORDER BY quarter
        """
        return pd.read_sql_query(query, conn)

    def analyze_sector_trends(self, start_year=2016, end_year=2024):
        """분야/부문별 연도별 추이 분석"""
        conn = self.connect_db()
        results = []
        
        for year in range(start_year, end_year + 1):
            query = f"""
            SELECT 
                '{year}' as year,
                fld_nm as sector,
                part_nm as subsector,
                COUNT(DISTINCT dbiz_cd) as project_count,
                SUM(bdg_cash_amt) as total_budget,
                SUM(ep_amt) as total_execution,
                SUM(cpl_amt) as allocated_budget,  -- 편성액 추가
                CASE 
                    WHEN SUM(bdg_cash_amt) > 0 
                    THEN CAST(SUM(ep_amt) AS FLOAT) / SUM(bdg_cash_amt)
                    ELSE 0 
                END as execution_rate
            FROM yearly_data_{year}
            WHERE fld_nm IS NOT NULL
            GROUP BY fld_nm, part_nm
            """
            try:
                year_data = pd.read_sql_query(query, conn)
                results.append(year_data)
            except Exception as e:
                print(f"Error processing year {year}: {e}")
                continue
        
        if results:
            return pd.concat(results, ignore_index=True)
        return pd.DataFrame()

    def analyze_detailed_projects(self, year):
        """세부사업 수준 분석"""
        conn = self.connect_db()
        query = f"""
        WITH project_summary AS (
            SELECT 
                dbiz_cd,
                dbiz_nm,
                fld_nm,
                part_nm,
                wa_laf_hg_nm,
                laf_hg_nm,
                acnt_dv_nm,
                SUM(bdg_cash_amt) as total_budget,
                SUM(ep_amt) as total_execution,
                SUM(cpl_amt) as allocated_budget,  -- 편성액 추가
                COUNT(*) as transaction_count
            FROM yearly_data_{year}
            GROUP BY dbiz_cd, dbiz_nm, fld_nm, part_nm, wa_laf_hg_nm, laf_hg_nm, acnt_dv_nm
        )
        SELECT 
            *,
            CASE 
                WHEN total_budget > 0 THEN CAST(total_execution AS FLOAT) / total_budget
                ELSE 0 
            END as execution_rate
        FROM project_summary
        ORDER BY total_budget DESC
        """
        return pd.read_sql_query(query, conn)

    def analyze_regional_execution(self, year):
        """지역별 집행 현황 분석"""
        conn = self.connect_db()
        query = f"""
        WITH monthly_regional AS (
            SELECT 
                strftime('%m', exe_ymd) as month,
                wa_laf_hg_nm as region,
                laf_hg_nm as municipality,
                SUM(bdg_cash_amt) as total_budget,
                SUM(ep_amt) as total_execution,
                SUM(cpl_amt) as allocated_budget,  -- 편성액 추가
                SUM(bdg_ntep) as national_budget,
                SUM(capep) as provincial_budget,
                SUM(sggep) as municipal_budget
            FROM yearly_data_{year}
            GROUP BY strftime('%m', exe_ymd), wa_laf_hg_nm, laf_hg_nm
        )
        SELECT 
            *,
            CASE 
                WHEN total_budget > 0 THEN CAST(total_execution AS FLOAT) / total_budget
                ELSE 0 
            END as execution_rate
        FROM monthly_regional
        ORDER BY month, region, municipality
        """
        return pd.read_sql_query(query, conn)

    def analyze_multi_year_trends(self, start_year=2016, end_year=2024):
        """연도별 통합 분석"""
        conn = self.connect_db()
        results = []
        
        for year in range(start_year, end_year + 1):
            query = f"""
            SELECT 
                '{year}' as year,
                part_nm,
                SUM(CASE WHEN strftime('%m', exe_ymd) = '12' THEN bdg_cash_amt ELSE 0 END) as total_budget, 
                SUM(ep_amt) as total_execution,
                SUM(CASE WHEN strftime('%m', exe_ymd) = '12' THEN cpl_amt ELSE 0 END) as allocated_budget 
            FROM yearly_data_{year}
            GROUP BY part_nm
            """
            try:
                year_data = pd.read_sql_query(query, conn)
                results.append(year_data)
            except Exception as e:
                print(f"Error processing year {year}: {e}")
                continue
        
        if results:
            return pd.concat(results, ignore_index=True)
        return pd.DataFrame()

    def analyze_monthly_by_dimension(self, year, dimension):
        """월별/차원별 분석"""
        conn = self.connect_db()
        valid_dimensions = {
            'sector': 'fld_nm',
            'region': 'wa_laf_hg_nm',
            'subsector': 'part_nm',
            'account': 'acnt_dv_nm',
            'municipality': 'laf_hg_nm'
        }
        
        if dimension not in valid_dimensions:
            raise ValueError(f"Invalid dimension. Valid options are: {list(valid_dimensions.keys())}")
            
        dimension_col = valid_dimensions[dimension]
        
        query = f"""
        SELECT 
            strftime('%m', exe_ymd) as month,
            {dimension_col} as dimension_value,
            SUM(bdg_cash_amt) as total_budget,
            SUM(ep_amt) as total_execution,
            SUM(cpl_amt) as allocated_budget,  -- 편성액 추가
            COUNT(*) as record_count
        FROM yearly_data_{year}
        GROUP BY month, {dimension_col}
        ORDER BY month, {dimension_col}
        """
        df = pd.read_sql_query(query, conn)
        df['execution_rate'] = df.apply(
            lambda x: x['total_execution'] / x['total_budget'] if x['total_budget'] > 0 else 0, 
            axis=1
        )
        return df

    def analyze_quarterly_by_dimension(self, year, dimension):
        """분기별/차원별 분석"""
        conn = self.connect_db()
        valid_dimensions = {
            'sector': 'fld_nm',
            'region': 'wa_laf_hg_nm',
            'subsector': 'part_nm',
            'account': 'acnt_dv_nm',
            'municipality': 'laf_hg_nm'
        }
        
        if dimension not in valid_dimensions:
            raise ValueError(f"Invalid dimension. Valid options are: {list(valid_dimensions.keys())}")
            
        dimension_col = valid_dimensions[dimension]
        
        query = f"""
        SELECT 
            CASE 
                WHEN CAST(strftime('%m', exe_ymd) AS INTEGER) BETWEEN 1 AND 3 THEN '1'
                WHEN CAST(strftime('%m', exe_ymd) AS INTEGER) BETWEEN 4 AND 6 THEN '2'
                WHEN CAST(strftime('%m', exe_ymd) AS INTEGER) BETWEEN 7 AND 9 THEN '3'
                ELSE '4'
            END as quarter,
            {dimension_col} as dimension_value,
            SUM(bdg_cash_amt) as total_budget,
            SUM(ep_amt) as total_execution,
            SUM(cpl_amt) as allocated_budget,  -- 편성액 추가
            COUNT(*) as record_count
        FROM yearly_data_{year}
        GROUP BY quarter, {dimension_col}
        ORDER BY quarter, {dimension_col}
        """
        df = pd.read_sql_query(query, conn)
        df['execution_rate'] = df.apply(
            lambda x: x['total_execution'] / x['total_budget'] if x['total_budget'] > 0 else 0, 
            axis=1
        )
        return df

    def analyze_three_dimensions(self, year, time_period='quarter'):
        """3차원 분석 (시간/지역/분야)"""
        conn = self.connect_db()
        time_clause = """
            CASE 
                WHEN CAST(strftime('%m', exe_ymd) AS INTEGER) BETWEEN 1 AND 3 THEN '1'
                WHEN CAST(strftime('%m', exe_ymd) AS INTEGER) BETWEEN 4 AND 6 THEN '2'
                WHEN CAST(strftime('%m', exe_ymd) AS INTEGER) BETWEEN 7 AND 9 THEN '3'
                ELSE '4'
            END
        """ if time_period == 'quarter' else """
            CASE 
                WHEN CAST(strftime('%m', exe_ymd) AS INTEGER) <= 6 THEN '상반기'
                ELSE '하반기'
            END
        """
        
        query = f"""
        SELECT 
            {time_clause} as time_period,
            wa_laf_hg_nm as region,
            fld_nm as sector,
            SUM(bdg_cash_amt) as total_budget,
            SUM(ep_amt) as total_execution,
            SUM(cpl_amt) as allocated_budget,  -- 편성액 추가
            COUNT(*) as record_count
        FROM yearly_data_{year}
        GROUP BY time_period, wa_laf_hg_nm, fld_nm
        ORDER BY time_period, wa_laf_hg_nm, fld_nm
        """
        df = pd.read_sql_query(query, conn)
        df['execution_rate'] = df.apply(
            lambda x: x['total_execution'] / x['total_budget'] if x['total_budget'] > 0 else 0, 
            axis=1
        )
        return df

    def validate_data_continuity(self, year):
        """월별 데이터의 연속성 검증"""
        conn = self.connect_db()
        query = f"""
        WITH monthly_data AS (
            SELECT 
                strftime('%m', exe_ymd) as month,
                SUM(ep_amt) as total_execution,
                SUM(cpl_amt) as allocated_budget,  -- 편성액 추가
                COUNT(*) as record_count
            FROM yearly_data_{year}
            GROUP BY strftime('%m', exe_ymd)
            ORDER BY month
        )
        SELECT 
            month,
            total_execution,
            allocated_budget,
            record_count,
            total_execution - LAG(total_execution) OVER (ORDER BY month) as monthly_change,
            CASE 
                WHEN total_execution < LAG(total_execution) OVER (ORDER BY month)
                THEN 1
                ELSE 0
            END as continuity_error
        FROM monthly_data
        """
        return pd.read_sql_query(query, conn)

    def detect_anomalies(self, year):
        """이상치 탐지 (12월 말 기준)"""
        conn = self.connect_db()
        try:
            query_data = f"""
            SELECT 
                dbiz_cd,
                dbiz_nm,
                fld_nm,
                part_nm,
                bdg_cash_amt,
                ep_amt,
                cpl_amt  -- 편성액 추가
            FROM yearly_data_{year}
            WHERE strftime('%m', exe_ymd) = '12'
            """
            
            df = pd.read_sql_query(query_data, conn)
            
            df['집행률'] = df.apply(
                lambda x: x['ep_amt'] / x['bdg_cash_amt'] if x['bdg_cash_amt'] > 0 else 0, 
                axis=1
            )
            
            df['이상유형'] = None
            df.loc[df['집행률'] > 1.0, '이상유형'] = '100% 초과집행'
            df.loc[(df['bdg_cash_amt'] > 0) & (df['ep_amt'] == 0), '이상유형'] = '미집행'
            df.loc[(df['bdg_cash_amt'] == 0) & (df['ep_amt'] > 0), '이상유형'] = '예산없는 집행'
            
            anomaly_df = df[df['이상유형'].notnull()].copy()
            
            anomaly_df = anomaly_df.rename(columns={
                'dbiz_cd': '사업코드',
                'dbiz_nm': '사업명',
                'fld_nm': '분야',
                'part_nm': '부문',
                'bdg_cash_amt': '예산액',
                'ep_amt': '집행액',
                'cpl_amt': '편성액'
            })
            
            anomaly_df = anomaly_df.sort_values(
                by=['이상유형', '집행률'], 
                ascending=[True, False]
            )
            
            return anomaly_df
            
        except Exception as e:
            print(f"이상치 탐지 과정에서 오류 발생: {e}")
            return pd.DataFrame(columns=[
                '사업코드', '사업명', '분야', '부문', '예산액', '집행액', '편성액', '집행률', '이상유형'
            ])

    def analyze_year_over_year(self, start_year=2016, end_year=2024):
        """전년 대비 증감률 분석"""
        conn = self.connect_db()
        results = []
        
        for year in range(start_year + 1, end_year + 1):
            query = f"""
            WITH current_year AS (
                SELECT 
                    '{year}' as year,
                    part_nm,
                    SUM(bdg_cash_amt) as current_budget,
                    SUM(ep_amt) as current_execution,
                    SUM(cpl_amt) as current_allocated  -- 편성액 추가
                FROM yearly_data_{year}
                GROUP BY part_nm
            ),
            previous_year AS (
                SELECT 
                    part_nm,
                    SUM(bdg_cash_amt) as prev_budget,
                    SUM(ep_amt) as prev_execution,
                    SUM(cpl_amt) as prev_allocated  -- 편성액 추가
                FROM yearly_data_{year-1}
                GROUP BY part_nm
            )
            SELECT 
                cy.year,
                cy.part_nm,
                cy.current_budget,
                cy.current_execution,
                cy.current_allocated,
                py.prev_budget,
                py.prev_execution,
                py.prev_allocated,
                CASE 
                    WHEN py.prev_budget > 0 
                    THEN (cy.current_budget - py.prev_budget) / py.prev_budget
                    ELSE NULL 
                END as budget_growth_rate,
                CASE 
                    WHEN py.prev_execution > 0 
                    THEN (cy.current_execution - py.prev_execution) / py.prev_execution
                    ELSE NULL 
                END as execution_growth_rate
            FROM current_year cy
            LEFT JOIN previous_year py ON cy.part_nm = py.part_nm
            """
            try:
                year_data = pd.read_sql_query(query, conn)
                results.append(year_data)
            except Exception as e:
                print(f"Error processing year {year}: {e}")
                continue
        
        if results:
            return pd.concat(results, ignore_index=True)
        return pd.DataFrame()

    def create_excel_report(self, output_path, year=None, multi_year=False):
        """분석 결과를 엑셀 파일로 저장"""
        wb = openpyxl.Workbook()
        
        if multi_year:
            multi_year_data = self.analyze_multi_year_trends()
            self._create_multi_year_sheet(wb, multi_year_data)
            
            growth_data = self.analyze_year_over_year()
            self._create_growth_analysis_sheet(wb, growth_data)
        
        if year:
            validation_data = self.validate_data_continuity(year)
            anomaly_data = self.detect_anomalies(year)
            self._create_validation_sheet(wb, validation_data, anomaly_data)
            
            seasonality_data = self.analyze_seasonality(year)
            self._create_seasonality_sheet(wb, seasonality_data)
            
            monthly_data = self.analyze_execution_by_month(year)
            self._create_monthly_sheet(wb, monthly_data)
            
            for dimension in ['sector', 'region']:
                monthly_dim_data = self.analyze_monthly_by_dimension(year, dimension)
                self._create_dimension_sheet(wb, monthly_dim_data, f'월별_{dimension}별_분석')
            
            for dimension in ['account', 'subsector']:
                quarterly_dim_data = self.analyze_quarterly_by_dimension(year, dimension)
                self._create_dimension_sheet(wb, quarterly_dim_data, f'분기별_{dimension}별_분석')
            
            three_dim_data = self.analyze_three_dimensions(year)
            self._create_three_dim_sheet(wb, three_dim_data)
            
            muni_data = self.analyze_monthly_by_dimension(year, 'municipality')
            self._create_dimension_sheet(wb, muni_data, '자치단체별_분석')
        
        self._apply_workbook_styles(wb)
        wb.save(output_path)

    def _create_multi_year_sheet(self, wb, df):
        """다년도 분석 시트 생성"""
        ws = wb.active
        ws.title = '다년도_분석'
        
        pivot = pd.pivot_table(
            df,
            values=['total_budget', 'total_execution', 'allocated_budget'],
            index=['part_nm'],
            columns=['year'],
            aggfunc='sum'
        )
        
        headers = ['부문'] + [f'{year}_예산' for year in range(2016, 2025)] + \
                 [f'{year}_지출' for year in range(2016, 2025)] + \
                 [f'{year}_편성' for year in range(2016, 2025)]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        
        for idx, (index, row) in enumerate(pivot.iterrows(), 2):
            ws.cell(row=idx, column=1, value=index)
            for year_idx, year in enumerate(range(2016, 2025)):
                ws.cell(row=idx, column=2+year_idx, value=row.get(('total_budget', str(year)), 0))
                ws.cell(row=idx, column=11+year_idx, value=row.get(('total_execution', str(year)), 0))
                ws.cell(row=idx, column=20+year_idx, value=row.get(('allocated_budget', str(year)), 0))

    def _create_dimension_sheet(self, wb, df, sheet_name):
        """차원별 분석 시트 생성"""
        ws = wb.create_sheet(sheet_name)
        
        headers = ['시간구분', '구분', '예산현액', '지출액', '편성액', '집행률(%)', '건수']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        
        for idx, row in df.iterrows():
            time_col = row['month'] if 'month' in df.columns else row['quarter']
            ws.cell(row=idx+2, column=1, value=time_col)
            ws.cell(row=idx+2, column=2, value=row['dimension_value'])
            ws.cell(row=idx+2, column=3, value=float(row['total_budget']))
            ws.cell(row=idx+2, column=4, value=float(row['total_execution']))
            ws.cell(row=idx+2, column=5, value=float(row['allocated_budget']))
            ws.cell(row=idx+2, column=6, value=float(row['execution_rate']))
            ws.cell(row=idx+2, column=7, value=int(row['record_count']))

    def _create_three_dim_sheet(self, wb, df):
        """3차원 분석 시트 생성"""
        ws = wb.create_sheet('3차원_분석')
        
        headers = ['시간구분', '지역', '분야', '예산현액', '지출액', '편성액', '집행률(%)', '건수']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        
        for idx, row in df.iterrows():
            ws.cell(row=idx+2, column=1, value=row['time_period'])
            ws.cell(row=idx+2, column=2, value=row['region'])
            ws.cell(row=idx+2, column=3, value=row['sector'])
            ws.cell(row=idx+2, column=4, value=float(row['total_budget']))
            ws.cell(row=idx+2, column=5, value=float(row['total_execution']))
            ws.cell(row=idx+2, column=6, value=float(row['allocated_budget']))
            ws.cell(row=idx+2, column=7, value=float(row['execution_rate']))
            ws.cell(row=idx+2, column=8, value=int(row['record_count']))

    def _create_monthly_sheet(self, wb, df):
        """월별 집행 현황 시트 생성"""
        ws = wb.active
        ws.title = '월별 집행 현황'
        
        headers = ['월', '예산현액', '지출액', '편성액', '집행률(%)', '국비', '시도비', 
                  '시군구비', '기타', '건수']
        
        header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True)
        
        for idx, row in df.iterrows():
            ws.cell(row=idx+2, column=1, value=int(row['month']))
            ws.cell(row=idx+2, column=2, value=float(row['total_budget']))
            ws.cell(row=idx+2, column=3, value=float(row['total_execution']))
            ws.cell(row=idx+2, column=4, value=float(row['allocated_budget']))
            ws.cell(row=idx+2, column=5, value=float(row['cumulative_execution_rate']) * 100)
            ws.cell(row=idx+2, column=6, value=float(row['national_budget']))
            ws.cell(row=idx+2, column=7, value=float(row['provincial_budget']))
            ws.cell(row=idx+2, column=8, value=float(row['municipal_budget']))
            ws.cell(row=idx+2, column=9, value=float(row['other_budget']))
            ws.cell(row=idx+2, column=10, value=int(row['record_count']))

        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        self._create_monthly_chart(ws, len(df) + 2)

    def _create_monthly_chart(self, ws, max_row):
        """월별 집행 현황 차트 생성"""
        line_chart = LineChart()
        line_chart.title = "월별 예산액과 지출액 추이"
        line_chart.style = 10
        line_chart.height = 10
        line_chart.width = 20
        
        cats = Reference(ws, min_col=1, min_row=2, max_row=max_row-1)
        data = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=max_row-1)  # 편성액 포함
        
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(cats)
        
        ws.add_chart(line_chart, "A{}".format(max_row + 2))
        
        bar_chart = BarChart()
        bar_chart.title = "월별 집행률"
        bar_chart.style = 10
        bar_chart.height = 10
        bar_chart.width = 20
        
        data = Reference(ws, min_col=5, max_col=5, min_row=1, max_row=max_row-1)
        
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)
        
        ws.add_chart(bar_chart, "J{}".format(max_row + 2))

    def _apply_workbook_styles(self, wb):
        """워크북 전체에 스타일 적용"""
        number_format = '#,##0'
        percent_format = '0.00%'
        
        for ws in wb.worksheets:
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2
            
            for row in ws.iter_rows(min_row=2):
                for cell in row[1:]:
                    if isinstance(cell.value, (int, float)):
                        if '집행률' in str(ws.cell(row=1, column=cell.column).value):
                            cell.number_format = percent_format
                        else:
                            cell.number_format = number_format

    def _create_growth_analysis_sheet(self, wb, df):
        """전년대비 증감률 분석 시트 생성"""
        ws = wb.create_sheet('전년대비_증감분석')
        
        headers = ['연도', '부문', '당해연도_예산', '당해연도_집행액', '당해연도_편성액',
                  '전년도_예산', '전년도_집행액', '전년도_편성액', '예산증감률(%)', '집행증감률(%)']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        
        for idx, row in df.iterrows():
            ws.cell(row=idx+2, column=1, value=row['year'])
            ws.cell(row=idx+2, column=2, value=row['part_nm'])
            ws.cell(row=idx+2, column=3, value=float(row['current_budget']))
            ws.cell(row=idx+2, column=4, value=float(row['current_execution']))
            ws.cell(row=idx+2, column=5, value=float(row['current_allocated']))
            ws.cell(row=idx+2, column=6, value=float(row['prev_budget']))
            ws.cell(row=idx+2, column=7, value=float(row['prev_execution']))
            ws.cell(row=idx+2, column=8, value=float(row['prev_allocated']))
            ws.cell(row=idx+2, column=9, value=float(row['budget_growth_rate'] or 0))
            ws.cell(row=idx+2, column=10, value=float(row['execution_growth_rate'] or 0))

    def _create_validation_sheet(self, wb, validation_df, anomaly_df):
        """데이터 검증 결과 시트 생성"""
        ws = wb.create_sheet('데이터_검증')
        
        ws.cell(row=1, column=1, value='월별 데이터 연속성 검증')
        headers = ['월', '총집행액', '편성액', '데이터건수', '전월대비증감', '연속성오류']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        
        for idx, row in validation_df.iterrows():
            ws.cell(row=idx+3, column=1, value=row['month'])
            ws.cell(row=idx+3, column=2, value=float(row['total_execution']))
            ws.cell(row=idx+3, column=3, value=float(row['allocated_budget']))
            ws.cell(row=idx+3, column=4, value=int(row['record_count']))
            ws.cell(row=idx+3, column=5, value=float(row['monthly_change'] or 0))
            ws.cell(row=idx+3, column=6, value=int(row['continuity_error']))
        
        start_row = len(validation_df) + 5
        ws.cell(row=start_row, column=1, value='이상치 목록')
        headers = ['사업코드', '사업명', '분야', '부문', '예산액', '집행액', '편성액', '집행률', '이상유형']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=col, value=header)
            cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        
        for idx, row in anomaly_df.iterrows():
            ws.cell(row=start_row+2+idx, column=1, value=row['사업코드'])
            ws.cell(row=start_row+2+idx, column=2, value=row['사업명'])
            ws.cell(row=start_row+2+idx, column=3, value=row['분야'])
            ws.cell(row=start_row+2+idx, column=4, value=row['부문'])
            ws.cell(row=start_row+2+idx, column=5, value=float(row['예산액']))
            ws.cell(row=start_row+2+idx, column=6, value=float(row['집행액']))
            ws.cell(row=start_row+2+idx, column=7, value=float(row['편성액']))
            ws.cell(row=start_row+2+idx, column=8, value=float(row['집행률']))
            ws.cell(row=start_row+2+idx, column=9, value=row['이상유형'])

    def _create_seasonality_sheet(self, wb, df):
        """계절성 분석 시트 생성"""
        ws = wb.create_sheet('계절성_분석')
        
        headers = ['분기', '분기별실집행액', '분기별비중(%)']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        
        for idx, row in df.iterrows():
            ws.cell(row=idx+2, column=1, value=row['quarter'])
            ws.cell(row=idx+2, column=2, value=float(row['quarterly_execution']))
            ws.cell(row=idx+2, column=3, value=float(row['quarter_ratio']))
        
        self._create_seasonality_chart(ws, len(df) + 2)

    def _create_seasonality_chart(self, ws, start_row):
        """계절성 분석 차트 생성"""
        chart = BarChart()
        chart.title = "분기별 집행 패턴"
        chart.style = 10
        chart.height = 10
        chart.width = 20
        
        data = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=5)
        cats = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=5)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f"A{start_row}")

def run_analysis(start_year=2016, end_year=2024, batch_size=3):
    """전체 분석 실행 (연도별 배치 처리)"""
    analyzer = LocalFinanceAnalyzer()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    try:
        multi_year_output = f'multi_year_analysis_{timestamp}.xlsx'
        print("다년도 통합 분석 수행 중...")
        analyzer.create_excel_report(multi_year_output, multi_year=True)
        print(f"다년도 분석 완료: {multi_year_output}")
        
        for year_batch_start in range(start_year, end_year + 1, batch_size):
            year_batch_end = min(year_batch_start + batch_size, end_year + 1)
            print(f"\n{year_batch_start}-{year_batch_end-1}년 배치 분석 시작")
            
            for year in range(year_batch_start, year_batch_end):
                output_path = f'detailed_analysis_{year}_{timestamp}.xlsx'
                print(f"{year}년도 상세 분석 중...")
                try:
                    analyzer.create_excel_report(output_path, year=year)
                    print(f"{year}년도 분석 완료: {output_path}")
                except Exception as e:
                    print(f"Error processing year {year}: {e}")
                    continue
                
            gc.collect()
    
    except Exception as e:
        print(f"Analysis failed: {e}")
    finally:
        analyzer.close_db()

if __name__ == "__main__":
    import gc
    run_analysis()